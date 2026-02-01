"""Report generation and export endpoints."""
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from typing import Optional
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from app.models.report import TaxReport

router = APIRouter()


class ReportRequest(BaseModel):
    """Request model for report generation."""
    country: str = Field(..., description="Country code")
    year: int = Field(..., description="Tax year")
    format: str = Field(default="excel", description="Report format: excel, pdf, json")


def generate_excel_report(report: TaxReport) -> BytesIO:
    """
    Generate Excel report from tax report.
    
    Creates multiple sheets:
    1. Summary - Tax summary
    2. Transactions - Detailed transaction list
    3. Audit Trail - Calculation details
    """
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary", 0)
    ws_summary.append(["Crypto Tax Report"])
    ws_summary.append(["Country", report.country])
    ws_summary.append(["Year", report.year])
    ws_summary.append(["Generated At", report.generated_at.isoformat()])
    ws_summary.append([])
    
    # Summary data
    ws_summary.append(["Metric", "Amount (EUR)"])
    ws_summary.append(["Total Gains", str(report.summary.total_gains_eur)])
    ws_summary.append(["Total Losses", str(report.summary.total_losses_eur)])
    ws_summary.append(["Net Gain/Loss", str(report.summary.net_gain_loss_eur)])
    ws_summary.append(["Staking Rewards", str(report.summary.staking_rewards_eur)])
    ws_summary.append(["Taxable Amount", str(report.summary.taxable_amount_eur)])
    ws_summary.append(["Transaction Count", report.summary.transaction_count])
    
    # Style summary header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws_summary[5]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Transactions sheet
    ws_tx = wb.create_sheet("Transactions", 1)
    headers = [
        "ID", "Timestamp", "Type", "Chain", "Source",
        "Token In", "Amount In", "Token Out", "Amount Out",
        "Price In (USD)", "Price Out (USD)", "Price In (EUR)", "Price Out (EUR)",
        "Cost Basis (EUR)", "Proceeds (EUR)", "Gain/Loss (EUR)",
        "Holding Period (Days)", "Fee (EUR)", "Audit Notes"
    ]
    ws_tx.append(headers)
    
    # Style transaction header
    for cell in ws_tx[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add transactions
    for tx in report.transactions:
        row = [
            tx.id,
            tx.timestamp.isoformat(),
            tx.type.value,
            tx.chain,
            tx.source,
            tx.token_in.symbol if tx.token_in else "",
            str(tx.amount_in) if tx.amount_in else "",
            tx.token_out.symbol if tx.token_out else "",
            str(tx.amount_out) if tx.amount_out else "",
            str(tx.price_in_usd) if tx.price_in_usd else "",
            str(tx.price_out_usd) if tx.price_out_usd else "",
            str(tx.price_in_eur) if tx.price_in_eur else "",
            str(tx.price_out_eur) if tx.price_out_eur else "",
            str(tx.cost_basis_eur) if tx.cost_basis_eur else "",
            str(tx.proceeds_eur) if tx.proceeds_eur else "",
            str(tx.gain_loss_eur) if tx.gain_loss_eur else "",
            tx.holding_period_days if tx.holding_period_days else "",
            str(tx.fee_eur) if tx.fee_eur else "",
            tx.audit_notes or ""
        ]
        ws_tx.append(row)
    
    # Auto-adjust column widths
    for column in ws_tx.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_tx.column_dimensions[column_letter].width = adjusted_width
    
    # Audit trail sheet
    if report.audit_trail:
        ws_audit = wb.create_sheet("Audit Trail", 2)
        ws_audit.append(["Audit Trail"])
        ws_audit.append([])
        for line in report.audit_trail.split("\n"):
            ws_audit.append([line])
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


@router.post("/generate")
async def generate_report(request: ReportRequest):
    """
    Generate a tax report in the specified format.
    
    Returns an Excel file with:
    - Summary sheet
    - Detailed transactions
    - Audit trail
    """
    # This endpoint should receive the tax report data
    # For now, we'll return an error if called directly
    # The actual report generation should be called from the tax calculation endpoint
    raise HTTPException(
        status_code=400,
        detail="Please use /api/v1/tax/calculate to generate a report first"
    )


@router.get("/download/{report_id}")
async def download_report(report_id: str, format: str = "excel"):
    """
    Download a previously generated report.
    
    Note: Without authentication, reports are not persisted.
    This endpoint will generate on-demand.
    """
    # TODO: Implement report storage and retrieval
    raise HTTPException(
        status_code=501,
        detail="Report download not yet implemented. Reports are generated on-demand."
    )
